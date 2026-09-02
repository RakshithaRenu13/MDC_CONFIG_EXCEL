import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from pathlib import Path
import openpyxl

# ============================================================
# MDC CONFIGURATOR
# Excel-driven Single Rack MDC database
#
# Put the Excel workbook in the same folder as this app:
# 1 Rack SKU'S - MDC BOQ (01.09.2026).xlsx
#
# The application reads the workbook every time Streamlit
# starts/reloads, so changes to the Excel database can flow
# into the UI without manually rewriting the BOM database.
# ============================================================

st.set_page_config(
    page_title="MDC Configuration & BOM Generator",
    page_icon="🏭",
    layout="wide"
)

EXCEL_FILENAME = "1 Rack SKU'S - MDC BOQ (01.09.2026).xlsx"
EXCEL_PATH = Path(__file__).resolve().parent / EXCEL_FILENAME

# ------------------------------------------------------------
# FALLBACK DATABASE GENERATED FROM THE PROVIDED EXCEL
# ------------------------------------------------------------
EMBEDDED_DATABASE = {'configs': {'Config 1': {'title': '', 'bom': []}, 'Config 3': {'title': 'SOLUTION 3  (7 kw Cooling W/o Dehumidifier)', 'bom': [{'part_number': 'CTO3M002', 'description': 'SINGLE RACK MDC, 7kW Cooling Unit, Without Dehumidifier', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801029209', 'description': 'MDC,42U 8*14 1R,INRACK7KW,EFSS', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'FRAME,42U-800X1200X2100  Front Glass Door , and rear sheet steel dual Spilt Door', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'FRONT DOOR WITH ELECTRONIC HANDLE WITH BIOMETRIC LOCK', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'REAR ELECTRONIC LOCK WITH EMERGENCY HIGH TEMPERATURE REAR DOOR OPENING', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'POWER DISTRIBUTION MODULE', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': '1 U SMART RACK MONITORING SYSTEM INTEGRATED WITH SMS, EMAIL - 1 NO.', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'TEMPERATURE SENSOR', 'quantity': 2, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'TEMPERATURE AND HUMIDITY SENSOR', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'WLD SENSOR CABLE', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'RODENT REPELLENT SYSTEM', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': '7” HMI DISPLAY', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'LED LIGHT, 12V LOGIC CONTROL', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'NORMAL LIGHTS WITH DOOR LIMIT SWITCH', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'BEACON ALARM, 12V, 108DB', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'BLANKING PANEL 1U', 'quantity': 20, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'MOUNTING HARDWARE PACK OF 20', 'quantity': 4, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'SMOKE DETECTOR', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'FULL TRAY', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401727', 'description': 'PAC, 7KW INVERTER RACK MNT WF ID', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401728', 'description': 'PAC, 7KW INVERTER RACK MNT WF OD', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401746', 'description': 'PAC, 7KW INV RACK MNT WF CONTROLLER', 'quantity': 1, 'uom': 'EA'}]}, 'Config 2': {'title': '', 'bom': []}, 'Config 4': {'title': 'SOLUTION 4 (7 kw Cooling With Dehumidifier)', 'bom': [{'part_number': 'CTO3M002', 'description': 'SINGLE RACK MDC, 7kW Cooling Unit, Dehumidifier', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801029209', 'description': 'MDC,42U 8*14 1R,INRACK7KW,EFSS', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801029209', 'description': 'FRAME,42U-800X1200X2100  Front Glass Door , and rear sheet steel dual Spilt Door', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'FRONT DOOR WITH ELECTRONIC HANDLE WITH BIOMETRIC LOCK', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'REAR ELECTRONIC LOCK WITH EMERGENCY HIGH TEMPERATURE REAR DOOR OPENING', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'POWER DISTRIBUTION MODULE', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': '1 U SMART RACK MONITORING SYSTEM INTEGRATED WITH SMS, EMAIL - 1 NO.', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'TEMPERATURE SENSOR', 'quantity': 2, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'TEMPERATURE AND HUMIDITY SENSOR', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'WLD SENSOR CABLE', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'RODENT REPELLENT SYSTEM', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': '7” HMI DISPLAY', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'LED LIGHT, 12V LOGIC CONTROL', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'NORMAL LIGHTS WITH DOOR LIMIT SWITCH', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'BEACON ALARM, 12V, 108DB', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'BLANKING PANEL 1U', 'quantity': 20, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'MOUNTING HARDWARE PACK OF 20', 'quantity': 4, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'SMOKE DETECTOR', 'quantity': 1, 'uom': 'EA'}, {'part_number': 'XXX', 'description': 'FULL TRAY', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401707', 'description': 'PAC, 8.6KW INV RACK MNT ATM S-007KAH', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401708', 'description': 'PAC, 12KW INV RACK MNT ATM OD KSF12AC', 'quantity': 1, 'uom': 'EA'}, {'part_number': '801401703', 'description': 'PAC, 3.5/7KW INVERTER RACK MNT ATM HMI', 'quantity': 1, 'uom': 'EA'}]}}, 'optional_items': [{'part_number': '801073203', 'description': 'FIRE SUPR EXT42U 8X14,1 RACK SOLN (EXTERNAL TYPE)', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': 'HRD-XH1C', 'description': 'FIRE SUPPRESS,RACK MNT FK-5-1-12,1.5m³', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303201', 'description': 'CAMERA,4MP VANDAL(CP-UNC-VC41L5C-VMD-LQ)', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303202', 'description': 'CAMERA,NVR 4 CHA INT(CP-UNR-4K4082-V4)', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303204', 'description': 'CAMERA,POE GB 4P,2UP(CP-DNW-GPU4G2-48C)', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303206', 'description': 'CAMERA,CAT 5 CABLE RJ45 TER', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303208', 'description': 'CAMERA,SVR HDD 1TB', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801303203', 'description': 'CAMERA,SVR HDD 4TB', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801223664', 'description': 'Rotating Keyboard tray', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801075237', 'description': '1 U CABLE MANAGER PLASTIC', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801029022', 'description': 'MDC,42U TOP CABLE TRAY IT', 'default_quantity': 1, 'uom': 'EA'}, {'part_number': '801075235', 'description': 'BRUSH PANEL 1 U', 'default_quantity': 1, 'uom': 'EA'}], 'pdu_items': [{'part_number': '802001004', 'description': 'B-PDU,ZU 20,32A 1P IEC309 20,0 230/230V', 'c13': 20, 'c19': 0, 'type': 'BASIC', 'uom': 'EA'}, {'part_number': '802001005', 'description': 'B-PDU,ZU 30,32A 1P IEC309 24,6 230/230V', 'c13': 24, 'c19': 6, 'type': 'BASIC', 'uom': 'EA'}, {'part_number': '802001006', 'description': 'B-PDU,ZU 42,32A 1P IEC309 36,6 230/230V', 'c13': 36, 'c19': 6, 'type': 'BASIC', 'uom': 'EA'}, {'part_number': '802001014', 'description': 'B-PDU,ZU 12,16A 1P IEC309 9,3 230/230V', 'c13': 9, 'c19': 3, 'type': 'BASIC', 'uom': 'EA'}, {'part_number': '802002003', 'description': 'M-PDU,ZU 24,16A 1P IEC309 20,4 230/230V', 'c13': 20, 'c19': 4, 'type': 'METERED', 'uom': 'EA'}, {'part_number': '802002005', 'description': 'M-PDU,ZU 24,32A 1P IEC309 20,4 230/230V', 'c13': 20, 'c19': 4, 'type': 'METERED', 'uom': 'EA'}, {'part_number': '802002006', 'description': 'M-PDU,ZU 34,32A 1P IEC309 28,6 230/230V', 'c13': 28, 'c19': 6, 'type': 'METERED', 'uom': 'EA'}, {'part_number': '802002007', 'description': 'M-PDU,ZU 42,32A 1P IEC309 36,6 230/230V', 'c13': 36, 'c19': 6, 'type': 'METERED', 'uom': 'EA'}, {'part_number': '802002012', 'description': 'M-PDU,ZU 34,32A 1P IEC309 28,6 230/230 W', 'c13': 28, 'c19': 6, 'type': 'METERED', 'uom': 'EA'}, {'part_number': '801601216', 'description': 'S-PDU ,ZU42,32A 1P IEC309 36,6 230/230V', 'c13': 36, 'c19': 6, 'type': 'SWITCHED', 'uom': 'EA'}, {'part_number': '802003002', 'description': 'S-PDU,ZU 16,16A 1P C20 12,4 230/230V', 'c13': 12, 'c19': 4, 'type': 'SWITCHED', 'uom': 'EA'}, {'part_number': '802003004', 'description': 'S-PDU,ZU 24,16A 1P IEC309 20,4 230/230V', 'c13': 20, 'c19': 4, 'type': 'SWITCHED', 'uom': 'EA'}, {'part_number': '802003006', 'description': 'S-PDU,ZU 24,32A 1P IEC309 20,4 230/400V', 'c13': 20, 'c19': 4, 'type': 'SWITCHED', 'uom': 'EA'}, {'part_number': '802003007', 'description': 'S-PDU,ZU 32,32A 1P IEC309 24,8 230/230V', 'c13': 24, 'c19': 8, 'type': 'SWITCHED', 'uom': 'EA'}, {'part_number': '802003008', 'description': 'S-PDU,ZU 44,32A 1P IEC309 34,6 230/230V', 'c13': 34, 'c19': 6, 'type': 'SWITCHED', 'uom': 'EA'}]}

# ============================================================
# EXCEL DATABASE READER
# ============================================================

def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def load_excel_database(path):
    """
    Reads the 1 Rack MDC BOQ workbook.

    Expected structure:
      - Solution 1: columns A:D
      - Solution 3: columns F:I
      - Solution 2: columns A:D
      - Solution 4: columns F:I
      - Other Optional Items: rows below the four solutions
      - Single Phase PDU catalogue: rows below optional items
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    solution_blocks = [
        (1, 1, "Config 1"),
        (1, 6, "Config 3"),
        (26, 1, "Config 2"),
        (26, 6, "Config 4"),
    ]

    configs = {}

    for title_row, start_col, config_name in solution_blocks:
        title = clean_text(ws.cell(title_row, start_col).value)
        bom = []

        row = title_row + 2

        while row <= ws.max_row:
            part = ws.cell(row, start_col).value
            description = ws.cell(row, start_col + 1).value
            quantity = ws.cell(row, start_col + 2).value
            uom = ws.cell(row, start_col + 3).value

            if all(v is None for v in (part, description, quantity, uom)):
                break

            if description is not None:
                bom.append({
                    "part_number": "XXX" if part is None else clean_text(part),
                    "description": clean_text(description),
                    "quantity": quantity if quantity is not None else 1,
                    "uom": "EA" if uom is None else clean_text(uom),
                })

            row += 1

        configs[config_name] = {
            "title": title,
            "bom": bom,
        }

    # Optional items from rows 52-63.
    optional_items = []

    for row in range(52, 64):
        part = ws.cell(row, 1).value
        description = ws.cell(row, 2).value
        quantity = ws.cell(row, 3).value
        uom = ws.cell(row, 4).value

        if description is not None:
            optional_items.append({
                "part_number": "XXX" if part is None else clean_text(part),
                "description": clean_text(description),
                "default_quantity": quantity if quantity is not None else 1,
                "uom": "EA" if uom is None else clean_text(uom),
            })

    # PDU catalogue from rows 66-80.
    pdu_items = []
    current_type = None

    for row in range(66, 81):
        part = ws.cell(row, 1).value
        description = ws.cell(row, 2).value
        c13 = ws.cell(row, 3).value
        c19 = ws.cell(row, 4).value
        item_type = ws.cell(row, 5).value

        if item_type:
            current_type = clean_text(item_type)

        if part is not None and description is not None:
            pdu_items.append({
                "part_number": clean_text(part),
                "description": clean_text(description),
                "c13": c13,
                "c19": c19,
                "type": current_type or "OTHER",
                "uom": "EA",
            })

    return {
        "configs": configs,
        "optional_items": optional_items,
        "pdu_items": pdu_items,
    }


def get_database():
    if EXCEL_PATH.exists():
        try:
            return load_excel_database(EXCEL_PATH), True, ""
        except Exception as exc:
            return EMBEDDED_DATABASE, False, str(exc)

    return EMBEDDED_DATABASE, False, (
        f"Excel database not found at: {EXCEL_PATH}. "
        "Using the database embedded from the supplied workbook."
    )


DATABASE, EXCEL_LOADED, EXCEL_MESSAGE = get_database()

if EXCEL_LOADED:
    st.sidebar.success("✅ Excel database loaded")
    st.sidebar.caption(EXCEL_FILENAME)
else:
    st.sidebar.warning("⚠️ Embedded database is being used")
    st.sidebar.caption(EXCEL_MESSAGE)

# ============================================================
# PRICE MASTER — EDIT ONLY IN CODE
#
# IMPORTANT:
# The supplied Excel contains BOM/SKU information, not prices.
# Therefore prices remain a separate code-only master.
# Replace XXX with official prices when available.
# ============================================================

STANDARD_PRICES = {
    "Config 1": 20.200,
    "Config 2": 30000,
    "Config 3": 27000,
    "Config 4": 32000,
}

OPTIONAL_PRICES = {
    "Fire Suppression System - External": 35000,
    "Fire Suppression System - In-Rack": 30000,
    "CAMERA,4MP VANDAL(CP-UNC-VC41L5C-VMD-LQ)": 34000,
    "UPS": "XXX",
    "Rotating Keyboard tray": "XXX",
    "1 U CABLE MANAGER PLASTIC": "XXX",
    "MDC,42U TOP CABLE TRAY IT": "XXX",
    "BRUSH PANEL 1 U": "XXX",
}

# ============================================================
# CURRENCY
# ============================================================

@st.cache_data(ttl=3600)
def get_live_exchange_rate(from_currency, to_currency):
    if from_currency == to_currency:
        return 1.0

    try:
        response = requests.get(
            f"https://api.frankfurter.dev/v2/rate/{from_currency}/{to_currency}",
            timeout=10
        )
        response.raise_for_status()
        return float(response.json()["rate"])
    except Exception:
        return None


def currency_symbol(currency):
    return "₹" if currency == "INR" else "$" if currency == "USD" else ""


def is_numeric_price(value):
    if isinstance(value, bool):
        return False

    if isinstance(value, (int, float)):
        return True

    if isinstance(value, str):
        try:
            float(value.strip().replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False

    return False


def numeric_price(value):
    if not is_numeric_price(value):
        return None
    return float(str(value).strip().replace(",", ""))


def format_price(value):
    number = numeric_price(value)

    if number is None:
        return "XXX"

    converted = number * exchange_rate
    return f"{currency_symbol(selected_currency)} {converted:,.2f}"


def calculate_amount(unit_price, quantity):
    number = numeric_price(unit_price)

    if number is None:
        return None

    return number * int(quantity)


# ============================================================
# CURRENCY SETTINGS
# ============================================================

st.header("💱 Currency Settings")

currency_col1, currency_col2 = st.columns([2, 2])

with currency_col1:
    selected_currency = st.radio(
        "Display Currency",
        ["INR", "USD"],
        horizontal=True
    )

if selected_currency == "INR":
    exchange_rate = 1.0
else:
    exchange_rate = get_live_exchange_rate("INR", "USD")

    if exchange_rate is None:
        st.error("Unable to fetch live INR → USD exchange rate.")
        st.stop()

if selected_currency == "USD":
    st.success(f"Live Exchange Rate: ₹1 = ${exchange_rate:.4f}")
else:
    st.info("Displaying all values in Indian Rupees (INR)")

st.divider()

# ============================================================
# HEADER
# ============================================================

st.title("🏭 MDC Configuration & BOM Generator")

st.markdown(
    """
    Select an MDC configuration, review the exact SKU/BOM data
    read from the Excel database, select optional items, and
    generate the commercial BOM.
    """
)

# ============================================================
# STEP 1 — MDC TYPE
# ============================================================

st.header("1️⃣ Select MDC Type")

mdc_type = st.selectbox(
    "MDC Type",
    ["Single Rack MDC"]
)

# ============================================================
# STEP 2 — CUSTOMER DETAILS
# ============================================================

st.header("2️⃣ Customer & Project Details")

col1, col2 = st.columns(2)

with col1:
    customer_name = st.text_input(
        "Customer Name *",
        placeholder="Enter customer / company name"
    )

with col2:
    customer_place = st.text_input(
        "Customer Place *",
        placeholder="Enter city / location"
    )

problem_col, solution_col = st.columns(2)

with problem_col:
    problem_statement = st.text_area(
        "Problem / Requirement",
        placeholder="Describe the customer's requirement..."
    )

with solution_col:
    proposed_solution = st.text_area(
        "Proposed Solution",
        placeholder="Describe the proposed MDC solution..."
    )

# ============================================================
# STEP 3 — CONFIGURATION
# ============================================================

st.header("3️⃣ Select Configuration")

config_data = DATABASE["configs"]

selected_config = st.selectbox(
    "Configuration",
    list(config_data.keys()),
    format_func=lambda key: f"{key} — {config_data[key]['title']}"
)

selected_config_data = config_data[selected_config]
selected_bom = selected_config_data["bom"]

# ============================================================
# STANDARD CONFIG PRICE
# ============================================================

standard_price = STANDARD_PRICES.get(selected_config, "XXX")

st.subheader("💵 Standard Configuration Cost")

p1, p2 = st.columns([2, 1])

with p1:
    st.metric(
        "Standard Cost",
        format_price(standard_price)
    )

with p2:
    st.caption(
        "Standard prices are maintained only in the Python PRICE MASTER. "
        "The Excel database supplies SKU/BOM information."
    )

# ============================================================
# STEP 4 — EXACT EXCEL BOM
# ============================================================

st.header("4️⃣ Complete Configuration BOM")

bom_preview_rows = []

for item in selected_bom:
    bom_preview_rows.append({
        "Part Number": item["part_number"],
        "Description": item["description"],
        "Quantity": item["quantity"],
        "UOM": item["uom"],
    })

bom_preview_df = pd.DataFrame(bom_preview_rows)

st.dataframe(
    bom_preview_df,
    use_container_width=True,
    hide_index=True
)

st.caption(
    f"{len(bom_preview_df)} BOM lines loaded from the Excel database."
)

# ============================================================
# STEP 5 — OPTIONAL ITEMS
# ============================================================

st.header("5️⃣ Optional Components")

st.info(
    "Optional items below are taken directly from the Excel "
    "section 'OTHER OPTIONAL ITEMS (TO BE INCLUDED AS PER NEED BASIS)'. "
    "Select an item and enter quantity."
)

selected_optional_items = []

for index, item in enumerate(DATABASE["optional_items"]):
    key_base = f"{selected_config}_{index}_{item['part_number']}_{item['description']}"

    c1, c2, c3, c4 = st.columns([0.7, 4.5, 1.2, 1.5])

    with c1:
        selected = st.checkbox(
            "",
            key=f"optional_selected_{key_base}"
        )

    with c2:
        st.write(f"**{item['description']}**")
        st.caption(f"Part Number: {item['part_number']}")

    with c3:
        unit_price = OPTIONAL_PRICES.get(item["description"], "XXX")
        st.write(f"Unit Cost: **{format_price(unit_price)}**")

    with c4:
        if selected:
            quantity = st.number_input(
                "Quantity",
                min_value=1,
                value=int(item["default_quantity"]) if str(item["default_quantity"]).isdigit() else 1,
                step=1,
                key=f"optional_qty_{key_base}"
            )
        else:
            quantity = 0

    if selected:
        unit_price = OPTIONAL_PRICES.get(item["description"], "XXX")
        amount = calculate_amount(unit_price, quantity)

        selected_optional_items.append({
            "part_number": item["part_number"],
            "description": item["description"],
            "quantity": quantity,
            "uom": item["uom"],
            "unit_price": unit_price,
            "amount": amount,
        })

# ============================================================
# PDU CATALOGUE
# ============================================================

st.subheader("🔌 Single Phase PDU Catalogue")

pdu_df = pd.DataFrame(DATABASE["pdu_items"])

if not pdu_df.empty:
    pdu_display = pdu_df[
        ["type", "part_number", "description", "c13", "c19", "uom"]
    ].rename(columns={
        "type": "Type",
        "part_number": "Part Number",
        "description": "Description",
        "c13": "C13",
        "c19": "C19",
        "uom": "UOM",
    })

    st.dataframe(
        pdu_display,
        use_container_width=True,
        hide_index=True
    )

    pdu_options = ["None"] + [
        f"{row['type']} | {row['part_number']} | {row['description']}"
        for _, row in pdu_df.iterrows()
    ]

    selected_pdu_label = st.selectbox(
        "Select PDU, if required",
        pdu_options
    )
else:
    selected_pdu_label = "None"

# Add selected PDU as an optional BOM line.
if selected_pdu_label != "None":
    selected_pdu = next(
        row for row in DATABASE["pdu_items"]
        if f"{row['type']} | {row['part_number']} | {row['description']}" == selected_pdu_label
    )

    pdu_qty = st.number_input(
        "PDU Quantity",
        min_value=1,
        value=1,
        step=1,
        key=f"pdu_quantity_{selected_config}"
    )

    pdu_price = OPTIONAL_PRICES.get(
        selected_pdu["description"],
        "XXX"
    )

    selected_optional_items.append({
        "part_number": selected_pdu["part_number"],
        "description": selected_pdu["description"],
        "quantity": pdu_qty,
        "uom": selected_pdu["uom"],
        "unit_price": pdu_price,
        "amount": calculate_amount(pdu_price, pdu_qty),
    })

# ============================================================
# OPTIONAL SUMMARY
# ============================================================

st.subheader("📋 Selected Optional Components")

if selected_optional_items:
    optional_rows = []

    for item in selected_optional_items:
        optional_rows.append({
            "Part Number": item["part_number"],
            "Description": item["description"],
            "Quantity": item["quantity"],
            "UOM": item["uom"],
            "Unit Cost": format_price(item["unit_price"]),
            "Amount": (
                format_price(item["amount"])
                if item["amount"] is not None
                else "XXX"
            ),
        })

    optional_df = pd.DataFrame(optional_rows)

    st.dataframe(
        optional_df,
        use_container_width=True,
        hide_index=True
    )
else:
    optional_df = pd.DataFrame(
        columns=[
            "Part Number",
            "Description",
            "Quantity",
            "UOM",
            "Unit Cost",
            "Amount",
        ]
    )

    st.info("No optional components selected.")

# ============================================================
# COST SUMMARY
# ============================================================

st.header("6️⃣ Cost Summary")

standard_numeric = numeric_price(standard_price)
optional_total = 0.0
unknown_optional_price = False

for item in selected_optional_items:
    if item["amount"] is None:
        unknown_optional_price = True
    else:
        optional_total += item["amount"]

if standard_numeric is not None and not unknown_optional_price:
    total_cost = standard_numeric + optional_total
else:
    total_cost = None

cost_rows = [{
    "Cost Item": "Standard Configuration",
    "Quantity": 1,
    "Unit Cost": format_price(standard_price),
    "Total Cost": format_price(standard_price),
}]

for item in selected_optional_items:
    cost_rows.append({
        "Cost Item": item["description"],
        "Quantity": item["quantity"],
        "Unit Cost": format_price(item["unit_price"]),
        "Total Cost": (
            format_price(item["amount"])
            if item["amount"] is not None
            else "XXX"
        ),
    })

cost_df = pd.DataFrame(cost_rows)

st.dataframe(
    cost_df,
    use_container_width=True,
    hide_index=True
)

c1, c2, c3 = st.columns(3)

with c1:
    st.metric(
        "Standard Configuration Cost",
        format_price(standard_price)
    )

with c2:
    st.metric(
        "Optional Components Cost",
        format_price(optional_total)
        if not unknown_optional_price
        else "XXX"
    )

with c3:
    st.metric(
        "TOTAL COST",
        format_price(total_cost)
        if total_cost is not None
        else "XXX"
    )

# ============================================================
# COST → PRICE
# ============================================================

st.subheader("📈 Cost → Price Conversion")

DEFAULT_PRICING_FACTORS = [
    ("Factory Cost (COGS)", 0.0),
    ("Admin & R&D Overhead", 15.0),
    ("Marketing & Sales", 20.0),
    ("Manufacturer Profit", 15.0),
    ("Distribution & Retail", 45.0),
]

pricing_factor_rows = []

for index, (default_name, default_percentage) in enumerate(
    DEFAULT_PRICING_FACTORS,
    start=1
):
    c1, c2 = st.columns([3, 1])

    with c1:
        factor_name = st.text_input(
            f"Layer {index} — Name",
            value=default_name,
            key=f"pricing_name_{index}"
        )

    with c2:
        factor_percentage = st.number_input(
            "Percentage",
            min_value=0.0,
            max_value=1000.0,
            value=default_percentage,
            step=0.5,
            format="%.2f",
            key=f"pricing_percentage_{index}"
        )

    pricing_factor_rows.append({
        "Layer": index,
        "Name": factor_name.strip(),
        "Percentage": float(factor_percentage),
    })

pricing_build_up_rows = []
running_amount = total_cost

for row in pricing_factor_rows:
    layer = row["Layer"]
    name = row["Name"]
    percentage = row["Percentage"]

    if layer == 1:
        previous_display = (
            format_price(total_cost)
            if total_cost is not None
            else "XXX"
        )
        added_display = "—"
        cumulative_display = (
            format_price(total_cost)
            if total_cost is not None
            else "XXX"
        )
    else:
        if running_amount is not None:
            previous_amount = running_amount
            added_amount = running_amount * (percentage / 100.0)
            running_amount += added_amount

            previous_display = format_price(previous_amount)
            added_display = f"+{percentage:.2f}% = {format_price(added_amount)}"
            cumulative_display = format_price(running_amount)
        else:
            previous_display = "XXX"
            added_display = f"+{percentage:.2f}% = XXX"
            cumulative_display = "XXX"

    pricing_build_up_rows.append({
        "Layer": layer,
        "Name": name,
        "Percentage Added": (
            "Baseline" if layer == 1 else f"{percentage:.2f}%"
        ),
        "Previous Amount": previous_display,
        "Added Amount": added_display,
        "Cumulative Price": cumulative_display,
    })

pricing_build_up_df = pd.DataFrame(pricing_build_up_rows)

st.dataframe(
    pricing_build_up_df,
    use_container_width=True,
    hide_index=True
)

selling_price = (
    running_amount
    if total_cost is not None
    else None
)

st.subheader("🏷️ Final Selling Price")

if selling_price is not None:
    st.success(f"### {format_price(selling_price)}")
else:
    st.success("### XXX")

# ============================================================
# WARNINGS
# ============================================================

if not customer_name.strip() or not customer_place.strip():
    st.warning("Customer Name and Customer Place are compulsory.")

if not is_numeric_price(standard_price):
    st.warning(
        f"{selected_config} standard price is XXX. "
        "Update STANDARD_PRICES in the Python PRICE MASTER."
    )

if unknown_optional_price:
    st.warning(
        "One or more selected optional items do not have a price. "
        "Update OPTIONAL_PRICES in the Python PRICE MASTER."
    )

# ============================================================
# STEP 7 — FINAL BOM
# ============================================================

st.header("7️⃣ Final BOM")

final_bom_rows = []

# Exact Excel BOM.
for item in selected_bom:
    final_bom_rows.append({
        "Category": "Standard",
        "Part Number": item["part_number"],
        "Description": item["description"],
        "Quantity": item["quantity"],
        "UOM": item["uom"],
        "Unit Price": "",
        "Amount": "",
    })

# Selected optional items.
for item in selected_optional_items:
    final_bom_rows.append({
        "Category": "Optional",
        "Part Number": item["part_number"],
        "Description": item["description"],
        "Quantity": item["quantity"],
        "UOM": item["uom"],
        "Unit Price": format_price(item["unit_price"]),
        "Amount": (
            format_price(item["amount"])
            if item["amount"] is not None
            else "XXX"
        ),
    })

final_bom_df = pd.DataFrame(
    final_bom_rows,
    columns=[
        "Category",
        "Part Number",
        "Description",
        "Quantity",
        "UOM",
        "Unit Price",
        "Amount",
    ]
)

st.dataframe(
    final_bom_df,
    use_container_width=True,
    hide_index=True
)

# ============================================================
# COMMERCIAL SUMMARY
# ============================================================

st.header("8️⃣ Final BOM & Commercial Summary")

commercial_rows = [{
    "Part Number": "XXX",
    "Description": f"Single Rack MDC — {selected_config}",
    "Quantity": 1,
    "Unit Price": format_price(standard_price),
    "Amount": format_price(standard_price),
}]

for item in selected_optional_items:
    commercial_rows.append({
        "Part Number": item["part_number"],
        "Description": item["description"],
        "Quantity": item["quantity"],
        "Unit Price": format_price(item["unit_price"]),
        "Amount": (
            format_price(item["amount"])
            if item["amount"] is not None
            else "XXX"
        ),
    })

commercial_df = pd.DataFrame(commercial_rows)

st.dataframe(
    commercial_df,
    use_container_width=True,
    hide_index=True
)

summary_df = pd.DataFrame([
    {"Metric": "Customer Name", "Value": customer_name or "—"},
    {"Metric": "Customer Place", "Value": customer_place or "—"},
    {"Metric": "MDC Type", "Value": mdc_type},
    {"Metric": "Selected Configuration", "Value": selected_config},
    {"Metric": "Problem / Requirement", "Value": problem_statement or "—"},
    {"Metric": "Proposed Solution", "Value": proposed_solution or "—"},
    {"Metric": "Standard Configuration Cost", "Value": format_price(standard_price)},
    {
        "Metric": "Optional Components Cost",
        "Value": (
            format_price(optional_total)
            if not unknown_optional_price
            else "XXX"
        ),
    },
    {
        "Metric": "TOTAL COST",
        "Value": format_price(total_cost) if total_cost is not None else "XXX",
    },
    {
        "Metric": "FINAL SELLING PRICE",
        "Value": format_price(selling_price) if selling_price is not None else "XXX",
    },
])

st.subheader("💰 Cost & Price Summary")

st.dataframe(
    summary_df,
    use_container_width=True,
    hide_index=True
)

# ============================================================
# EXCEL EXPORT
# ============================================================

st.header("9️⃣ Excel Export")

customer_details_df = pd.DataFrame([
    {"Field": "Customer Name", "Value": customer_name},
    {"Field": "Customer Place", "Value": customer_place},
    {"Field": "MDC Type", "Value": mdc_type},
    {"Field": "Selected Configuration", "Value": selected_config},
    {"Field": "Configuration Description", "Value": selected_config_data["title"]},
    {"Field": "Problem / Requirement", "Value": problem_statement},
    {"Field": "Proposed Solution", "Value": proposed_solution},
])

bom_without_price_df = final_bom_df[
    ["Category", "Part Number", "Description", "Quantity", "UOM"]
].copy()

bom_with_price_df = final_bom_df.copy()

price_summary_rows = [
    {"Section": "Customer", "Item": "Customer Name", "Value": customer_name},
    {"Section": "Customer", "Item": "Customer Place", "Value": customer_place},
    {"Section": "Configuration", "Item": "MDC Type", "Value": mdc_type},
    {"Section": "Configuration", "Item": "Selected Configuration", "Value": selected_config},
    {"Section": "Configuration", "Item": "Description", "Value": selected_config_data["title"]},
    {"Section": "Cost", "Item": "Standard Configuration Cost", "Value": format_price(standard_price)},
    {
        "Section": "Cost",
        "Item": "Optional Components Cost",
        "Value": format_price(optional_total) if not unknown_optional_price else "XXX",
    },
    {
        "Section": "Cost",
        "Item": "TOTAL COST",
        "Value": format_price(total_cost) if total_cost is not None else "XXX",
    },
    {
        "Section": "Price",
        "Item": "FINAL SELLING PRICE",
        "Value": format_price(selling_price) if selling_price is not None else "XXX",
    },
]

pricing_factors_export_df = pd.DataFrame(pricing_factor_rows)
pricing_factors_export_df["Percentage Added"] = pricing_factors_export_df[
    "Percentage"
].map(lambda x: "Baseline" if x == 0 else f"{x:.2f}%")
pricing_factors_export_df = pricing_factors_export_df[
    ["Layer", "Name", "Percentage Added"]
]

def create_excel_file(dataframes):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in dataframes.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False
            )

    output.seek(0)
    return output

excel_without_price = create_excel_file({
    "Customer Details": customer_details_df,
    "BOM": bom_without_price_df,
})

excel_with_price = create_excel_file({
    "Customer Details": customer_details_df,
    "BOM With Price": bom_with_price_df,
    "Cost Summary": cost_df,
    "Price Build-up": pricing_build_up_df,
    "Pricing Factors": pricing_factors_export_df,
    "Price Summary": pd.DataFrame(price_summary_rows),
})

download_col1, download_col2 = st.columns(2)

with download_col1:
    st.download_button(
        "📥 Download BOM — Without Price",
        data=excel_without_price,
        file_name=f"Single_Rack_{selected_config}_BOM_Without_Price.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_without_price",
    )

with download_col2:
    st.download_button(
        "📥 Download BOM — With Price",
        data=excel_with_price,
        file_name=f"Single_Rack_{selected_config}_BOM_With_Price.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_with_price",
    )

# ============================================================
# FOOTER
# ============================================================

st.divider()

st.caption(
    "MDC Configuration & BOM Generator | "
    "Single Rack SKU/BOM database is driven by the supplied Excel workbook. "
    "Prices remain separate in the Python PRICE MASTER."
)
